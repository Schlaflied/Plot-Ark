"""
Excel Report Exporter — generates multi-sheet Excel workbook with raw data + summaries.

Uses openpyxl. Reads from Postgres for raw xAPI data and feedback.
"""

import io

from services.chart_generator import COLORS


def export_excel(report: dict, course_id: int = 0) -> bytes:
    """Generate multi-sheet Excel with raw data + summaries."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return b"openpyxl not installed"

    from db import get_db

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="44403C")  # stone_700
    red_fill = PatternFill("solid", fgColor="FCA5A5")
    yellow_fill = PatternFill("solid", fgColor="FDE68A")

    # ── Sheet 1: Raw xAPI Data ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Raw xAPI Data"
    raw_headers = ["Student Name", "Email", "Verb", "Object ID", "Object Name", "Timestamp", "Course"]
    for c, h in enumerate(raw_headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill

    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT actor_name, actor_email, verb, object_id, object_name, timestamp, curriculum_topic
                FROM xapi_statements
                WHERE object_id LIKE %s
                ORDER BY timestamp
            """, (f"course/{course_id}/%",))
            for r_idx, row in enumerate(cur.fetchall(), 2):
                for c_idx, val in enumerate(row, 1):
                    ws1.cell(row=r_idx, column=c_idx,
                             value=val.isoformat() if hasattr(val, "isoformat") else val)
            cur.close()
            conn.close()
        except Exception as e:
            print(f"Excel raw data error: {e}")
            if conn:
                conn.close()

    # ── Sheet 2: Module Summary ───────────────────────────────────────────
    ws2 = wb.create_sheet("Module Summary")
    ba = report.get("behavior_analysis", {})
    mod_headers = ["Module", "Students", "Completions", "Struggles",
                   "Completion Rate", "Struggle Rate", "Drop-off Rate"]
    for c, h in enumerate(mod_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for r_idx, m in enumerate(ba.get("module_engagement", []), 2):
        ws2.cell(row=r_idx, column=1, value=m.get("module_name", ""))
        ws2.cell(row=r_idx, column=2, value=m.get("unique_students", 0))
        ws2.cell(row=r_idx, column=3, value=m.get("completions", 0))
        ws2.cell(row=r_idx, column=4, value=m.get("struggles", 0))
        ws2.cell(row=r_idx, column=5, value=min(m.get("completion_rate", 0), 1.0))
        ws2.cell(row=r_idx, column=6, value=m.get("struggle_rate", 0))
        ws2.cell(row=r_idx, column=7, value=m.get("drop_off_rate", 0))

    # ── Sheet 3: Student Roster ───────────────────────────────────────────
    ws3 = wb.create_sheet("Student Roster")
    stu_headers = ["Name", "Email", "Risk Level", "Risk Score",
                   "Total Actions", "Mastered", "Struggled", "Failed", "Completion Rate"]
    for c, h in enumerate(stu_headers, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill

    ra = report.get("risk_assessment", {})
    for r_idx, s in enumerate(ra.get("at_risk_students", []), 2):
        ws3.cell(row=r_idx, column=1, value=s.get("name", ""))
        ws3.cell(row=r_idx, column=2, value=s.get("email", ""))
        risk_cell = ws3.cell(row=r_idx, column=3, value=s.get("risk_level", ""))
        ws3.cell(row=r_idx, column=4, value=s.get("risk_score", 0))
        stats = s.get("stats", {})
        ws3.cell(row=r_idx, column=5, value=stats.get("total_actions", 0))
        ws3.cell(row=r_idx, column=6, value=stats.get("mastered", 0))
        ws3.cell(row=r_idx, column=7, value=stats.get("struggled", 0))
        ws3.cell(row=r_idx, column=8, value=stats.get("failed", 0))
        ws3.cell(row=r_idx, column=9, value=stats.get("completion_rate", 0))

        # Conditional formatting
        if s.get("risk_level") == "high":
            risk_cell.fill = red_fill
        elif s.get("risk_level") == "medium":
            risk_cell.fill = yellow_fill

    # ── Sheet 4: Feedback Summary ─────────────────────────────────────────
    ws4 = wb.create_sheet("Feedback Summary")
    fb_headers = ["Module", "Got It", "Mostly", "Something Off", "Didn't Read"]
    for c, h in enumerate(fb_headers, 1):
        cell = ws4.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill

    # Get feedback from DB
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT module_index, module_title,
                    COUNT(*) FILTER (WHERE sentiment = 'got-it') as got_it,
                    COUNT(*) FILTER (WHERE sentiment = 'mostly') as mostly,
                    COUNT(*) FILTER (WHERE sentiment = 'off') as off,
                    COUNT(*) FILTER (WHERE sentiment = 'not-read') as not_read
                FROM student_feedback
                WHERE course_id = %s
                GROUP BY module_index, module_title
                ORDER BY module_index
            """, (course_id,))
            for r_idx, row in enumerate(cur.fetchall(), 2):
                ws4.cell(row=r_idx, column=1, value=row[1] or f"Module {row[0]}")
                ws4.cell(row=r_idx, column=2, value=row[2])
                ws4.cell(row=r_idx, column=3, value=row[3])
                ws4.cell(row=r_idx, column=4, value=row[4])
                ws4.cell(row=r_idx, column=5, value=row[5])
            cur.close()
            conn.close()
        except Exception as e:
            print(f"Excel feedback error: {e}")
            if conn:
                conn.close()

    # Auto-width columns
    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
